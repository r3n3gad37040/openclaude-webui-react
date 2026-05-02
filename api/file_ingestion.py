"""File ingestion — extract text from common document formats."""
from __future__ import annotations

import json
from pathlib import Path


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".bmp", ".ico"}
TEXT_EXTS = {
    ".txt", ".md", ".markdown", ".csv", ".json", ".xml",
    ".py", ".js", ".ts", ".jsx", ".tsx", ".css", ".scss", ".less",
    ".html", ".htm", ".yaml", ".yml", ".toml", ".ini", ".cfg",
    ".sh", ".bash", ".zsh", ".fish", ".ps1", ".bat", ".cmd",
    ".c", ".cpp", ".h", ".hpp", ".rs", ".go", ".java", ".kt",
    ".rb", ".php", ".pl", ".r", ".swift", ".scala", ".clj",
    ".sql", ".dockerfile", ".makefile", ".cmake", ".gradle",
    ".log", ".conf", ".properties", ".env", ".gitignore",
    ".vue", ".svelte", ".astro", ".elm", ".clojure",
}


def is_image(path: Path) -> bool:
    return path.suffix.lower() in IMAGE_EXTS


def _read_text(path: Path) -> str:
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return path.read_bytes().decode("utf-8", errors="replace")


def _extract_docx(path: Path) -> str:
    try:
        import docx
        doc = docx.Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except Exception:
        return "[Could not extract DOCX content]"


def _extract_pdf(path: Path) -> str:
    try:
        import PyPDF2
        text = []
        with open(path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
        return "\n\n".join(text) if text else "[PDF contained no extractable text]"
    except Exception:
        return "[Could not extract PDF content]"


def _extract_xlsx(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            lines.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    lines.append(row_text)
            lines.append("")
        return "\n".join(lines)
    except Exception:
        return "[Could not extract XLSX content]"


def _extract_html(path: Path) -> str:
    try:
        from bs4 import BeautifulSoup
        html = _read_text(path)
        soup = BeautifulSoup(html, "html.parser")
        # Remove script and style elements
        for tag in soup(["script", "style", "nav", "footer"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
        # Collapse multiple blank lines
        lines = [line.strip() for line in text.splitlines()]
        return "\n".join(line for line in lines if line)
    except Exception:
        return _read_text(path)


def _extract_md(path: Path) -> str:
    try:
        import markdown
        md_text = _read_text(path)
        html = markdown.markdown(md_text)
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
        text = soup.get_text(separator="\n", strip=True)
        lines = [line.strip() for line in text.splitlines()]
        return "\n".join(line for line in lines if line)
    except Exception:
        return _read_text(path)


def extract_text(path: Path) -> str:
    """Extract plain text from a file path. Returns empty string for images."""
    p = Path(path)
    ext = p.suffix.lower()

    if ext in IMAGE_EXTS:
        return ""

    if ext in (".docx", ".doc"):
        return _extract_docx(p)

    if ext == ".pdf":
        return _extract_pdf(p)

    if ext in (".xlsx", ".xls"):
        return _extract_xlsx(p)

    if ext in (".html", ".htm"):
        return _extract_html(p)

    if ext in (".md", ".markdown"):
        return _extract_md(p)

    # Default: treat as plain text
    return _read_text(p)


def build_attachment_prompt(files: list[dict]) -> str:
    """Build a prompt block from uploaded file metadata.

    Only includes text-extractable files. Images are omitted from the
    prompt because the text-only CLI backend cannot process them.

    files: list of dicts with keys: name, path, size, extracted_text (optional)
    """
    parts = []
    texts = []

    for f in files:
        p = Path(f["path"])
        if not is_image(p):
            texts.append(f)

    if not texts:
        return ""

    parts.append("[Attached Documents]")
    for f in texts:
        text = f.get("extracted_text", "").strip()
        if not text:
            text = "[No extractable text]"
        parts.append(f"\n--- {f['name']} ---\n{text}")
    parts.append("\n[End of attached documents]\n")

    return "\n".join(parts)
